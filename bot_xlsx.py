import smtplib
from pathlib import Path
import openpyxl
import mimetypes
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from bot_create import email_login, email_password, email_server, admin_email
from sql import admin_requests


async def send_email(file_name):
    msg = MIMEMultipart()
    msg['From'] = email_login
    msg['To'] = admin_email
    msg['Subject'] = "Отчёт от бота"
    body = "Привет! Отчёт во вложении"
    msg.attach(MIMEText(body, 'plain'))
    attachment = Path(file_name)
    await attach_file(msg, attachment)
    server = smtplib.SMTP_SSL(email_server, 465)
    server.login(email_login, email_password)
    server.send_message(msg)
    server.quit()


async def attach_file(msg, attachment):
    filename = attachment.resolve(strict=True).name
    ctype, encoding = mimetypes.guess_type(attachment)
    if ctype is None or encoding is not None:
        ctype = 'application/octet-stream'
    maintype, subtype = ctype.split('/', 1)
    with open(attachment.resolve(strict=True), 'rb') as fp:
        file = MIMEBase(maintype, subtype)
        file.set_payload(fp.read())
        fp.close()
        encoders.encode_base64(file)
    file.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(file)


async def make_excel(table_name, file_name, column='row_id', param='%'):
    book = openpyxl.Workbook()
    sheet = book.active
    column_name = await admin_requests.sql_make_excel(table_name, column, param)
    data = await admin_requests.sql_make_excel(table_name, column, param)
    excel_column = 1
    for i in column_name[0]:
        sheet.cell(row=1, column=excel_column).value = i
        excel_column += 1
    row = 2
    for line in data[1]:
        excel_column2 = 1
        for i in line:
            sheet.cell(row=row, column=excel_column2).value = i
            excel_column2 += 1
        row += 1
    book.save(file_name)
    book.close()
